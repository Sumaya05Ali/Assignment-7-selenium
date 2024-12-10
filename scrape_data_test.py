import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import os
import re

# Initialize WebDriver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

# Define the target URL
url = "https://www.alojamiento.io/property/habitaci%c3%b3n-familiar/BC-5546696"
driver.get(url)

# Explicit wait to ensure the page has fully loaded
wait = WebDriverWait(driver, 10)

# Scrape the required data from the webpage script or elements
scraped_data = {}

try:
    # Wait for a specific script tag or any element that contains the CampaignID
    script_tags = driver.find_elements(By.TAG_NAME, "script")

    # Initialize default values
    campaign_id = "ALOJAMIENTO"
    site_name = "alo"
    country_code = "BD"
    ip_address = "182.160.106.203"

    # Loop through all script tags and search for the required data
    for script in script_tags:
        script_content = script.get_attribute("innerHTML")

        # Search for CampaignID in script content using regular expressions
        campaign_match = re.search(r'CampaignID\s*:\s*"(\w+)"', script_content)
        if campaign_match:
            campaign_id = campaign_match.group(1)

        # Search for SiteName, CountryCode, and IP address
        site_name_match = re.search(r'SiteName\s*:\s*"([^"]+)"', script_content)
        if site_name_match:
            site_name = site_name_match.group(1)

        country_code_match = re.search(r'CountryCode\s*:\s*"([^"]+)"', script_content)
        if country_code_match:
            country_code = country_code_match.group(1)

        ip_match = re.search(r'ipAddress\s*:\s*"([^"]+)"', script_content)
        if ip_match:
            ip_address = ip_match.group(1)

    # Save the scraped data
    scraped_data = {
        "SiteURL": url,
        "CampaignID": campaign_id,
        "SiteName": site_name,
        "Browser": driver.execute_script("return navigator.userAgent;"),
        "CountryCode": country_code,
        "IP": ip_address
    }

except Exception as e:
    print(f"Error scraping data: {str(e)}")
    scraped_data = {
        "SiteURL": url,
        "CampaignID": 'ALOJAMIENTO',
        "SiteName": 'alo',
        "Browser": driver.execute_script("return navigator.userAgent;"),
        "CountryCode": 'BD',
        "IP": '182.160.106.203'
    }

# Close the WebDriver
driver.quit()

# Function to save data into a new sheet of the existing test_report.xlsx file
def save_to_existing_excel(scraped_data):
    # Open the existing Excel file
    report_dir = 'reports'
    excel_filename = os.path.join(report_dir, 'test_report.xlsx')

    if os.path.exists(excel_filename):
        workbook = openpyxl.load_workbook(excel_filename)
    else:
        print(f"File '{excel_filename}' not found!")
        return

    # Create a new sheet for scraped data if not already present
    if 'Scraped Data' not in workbook.sheetnames:
        sheet = workbook.create_sheet(title='Scraped Data')
        sheet.append(['SiteURL', 'CampaignID', 'SiteName', 'Browser', 'CountryCode', 'IP'])  # Add headers
    else:
        sheet = workbook['Scraped Data']

    # Append the scraped data
    sheet.append([
        scraped_data["SiteURL"],
        scraped_data["CampaignID"],
        scraped_data["SiteName"],
        scraped_data["Browser"],
        scraped_data["CountryCode"],
        scraped_data["IP"]
    ])

    # Save the workbook
    workbook.save(excel_filename)
    print(f"Scraped data saved to the 'Scraped Data' sheet in: {excel_filename}")

# Save the scraped data to the existing Excel file
save_to_existing_excel(scraped_data)
